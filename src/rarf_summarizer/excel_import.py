from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rarf_summarizer.schema import Schema
from rarf_summarizer.storage import Store


def sync_back(store: Store, schema: Schema, path: Path) -> int:
    wb = load_workbook(path, data_only=True)
    ws = wb["RARF Overview"]
    headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    try:
        paper_col = headers.index("Paper ID") + 1
    except ValueError as exc:
        raise ValueError("RARF Overview is missing a Paper ID column") from exc
    label_to_id = {spec.label: spec.id for spec in schema.fields}
    field_cols = []
    for col, header in enumerate(headers, start=1):
        if header in label_to_id:
            field_cols.append((col, label_to_id[header]))
    updates = 0
    for row in range(3, ws.max_row + 1):
        paper_id = ws.cell(row, paper_col).value
        if not paper_id:
            continue
        for col, field_id in field_cols:
            value = ws.cell(row, col).value
            text = "" if value is None else str(value)
            existing = store.get_field(str(paper_id), field_id)
            last = (existing or {}).get("last_exported_text") or ""
            if text.strip() != str(last).strip():
                store.set_human_override(str(paper_id), field_id, text)
                updates += 1
    return updates
