from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from rarf_summarizer.formatting import effective_text
from rarf_summarizer.schema import Schema
from rarf_summarizer.storage import Store

ARIAL = Font(name="Arial", size=10)
ARIAL_BOLD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
GROUP_FILLS = {
    "identity": PatternFill("solid", fgColor="1F4E79"),
    "theory": PatternFill("solid", fgColor="196F3D"),
    "constructs": PatternFill("solid", fgColor="7D3C98"),
    "method": PatternFill("solid", fgColor="B7950B"),
    "implications": PatternFill("solid", fgColor="922B21"),
}
COL_FILLS = {
    "identity": PatternFill("solid", fgColor="D6EAF8"),
    "theory": PatternFill("solid", fgColor="D5F5E3"),
    "constructs": PatternFill("solid", fgColor="E8DAEF"),
    "method": PatternFill("solid", fgColor="FCF3CF"),
    "implications": PatternFill("solid", fgColor="FADBD8"),
}
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

IDENTITY_COLUMNS = [
    ("paper_id", "Paper ID"),
    ("title", "Title"),
    ("authors", "Authors"),
    ("year", "Year"),
    ("folder", "Folder"),
    ("relative_path", "Relative path"),
]


def export_workbook(store: Store, schema: Schema, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _overview(wb, store, schema)
    _variable_map(wb, store)
    _evidence(wb, store, schema)
    _run_log(wb, store)
    _notes(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)
    return path


def _style_header(cell, fill: PatternFill) -> None:
    cell.font = ARIAL_BOLD
    cell.fill = fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = THIN


def _style_body(cell, fill: PatternFill | None = None) -> None:
    cell.font = ARIAL
    cell.alignment = WRAP
    cell.border = THIN
    if fill:
        cell.fill = fill


def _overview(wb: Workbook, store: Store, schema: Schema) -> None:
    ws = wb.active
    ws.title = "RARF Overview"
    identity = IDENTITY_COLUMNS
    fields = list(schema.fields)
    groups: list[tuple[str, int]] = [("Identity", len(identity))]
    current_group = fields[0].group if fields else "theory"
    count = 0
    for spec in fields:
        if spec.group != current_group:
            groups.append((current_group.title(), count))
            current_group = spec.group
            count = 0
        count += 1
    if fields:
        groups.append((current_group.title(), count))

    col = 1
    for label, width in groups:
        start = col
        end = col + width - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, label)
        group_key = label.casefold().split()[0]
        _style_header(cell, GROUP_FILLS.get(group_key, HEADER_FILL))
        col = end + 1

    headers = [label for _, label in identity] + [spec.label for spec in fields]
    fills = ["identity"] * len(identity) + [spec.group for spec in fields]
    for index, (header, group) in enumerate(zip(headers, fills), start=1):
        cell = ws.cell(2, index, header)
        _style_header(cell, GROUP_FILLS.get(group, HEADER_FILL))

    papers = store.list_papers()
    for row_index, paper in enumerate(papers, start=3):
        values = [
            paper.get("id"),
            paper.get("title"),
            paper.get("authors"),
            paper.get("year"),
            paper.get("folder"),
            paper.get("relative_path"),
        ]
        fields_map = store.fields_for(paper["id"])
        exported = {}
        for spec in fields:
            text = effective_text(fields_map.get(spec.id))
            values.append(text)
            exported[spec.id] = text
            store.mark_exported(paper["id"], spec.id, text)
        for col_index, (value, group) in enumerate(zip(values, fills), start=1):
            cell = ws.cell(row_index, col_index, value)
            _style_body(cell, COL_FILLS.get(group))
        ws.row_dimensions[row_index].height = 90

    last_col = len(headers)
    last_row = max(2, 2 + len(papers))
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "G3"
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{max(last_row, 3)}"
    widths = [22, 36, 28, 10, 22, 40] + [28] * len(fields)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.oddHeader.left.text = "RARF Overview"
    ws.oddFooter.right.text = "Page &P of &N"


def _variable_map(wb: Workbook, store: Store) -> None:
    ws = wb.create_sheet("Variable-Measure Map")
    headers = [
        "Paper ID",
        "Construct ID",
        "Class",
        "Construct name",
        "Nominal definition",
        "Measure name",
        "Operationalization",
        "Range",
        "Type",
        "Linked construct",
    ]
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(1, col, header), HEADER_FILL)
    constructs = {(row["paper_id"], row["construct_id"]): row for row in store.list_constructs()}
    measures = store.list_measures()
    used = set()
    row_index = 2
    for measure in measures:
        key = (measure["paper_id"], measure.get("construct_id"))
        construct = constructs.get(key, {})
        used.add(key)
        values = [
            measure["paper_id"],
            measure.get("construct_id"),
            construct.get("class") or measure.get("class"),
            construct.get("name"),
            construct.get("nominal_definition"),
            measure.get("name"),
            measure.get("operationalization"),
            measure.get("range"),
            measure.get("type"),
            measure.get("linked_construct"),
        ]
        for col, value in enumerate(values, start=1):
            _style_body(ws.cell(row_index, col, value), COL_FILLS["constructs"])
        row_index += 1
    for key, construct in constructs.items():
        if key in used:
            continue
        values = [
            construct["paper_id"],
            construct["construct_id"],
            construct.get("class"),
            construct.get("name"),
            construct.get("nominal_definition"),
            None,
            None,
            None,
            None,
            None,
        ]
        for col, value in enumerate(values, start=1):
            _style_body(ws.cell(row_index, col, value), COL_FILLS["theory"])
        row_index += 1
    for index, width in enumerate([22, 18, 12, 24, 40, 22, 40, 14, 14, 22], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{max(row_index - 1, 1)}"


def _evidence(wb: Workbook, store: Store, schema: Schema) -> None:
    ws = wb.create_sheet("Evidence & QA")
    headers = [
        "Paper ID",
        "Field",
        "Status",
        "Confidence",
        "Page",
        "Quote",
        "Quote matched",
        "Match score",
        "Academic paraphrase",
        "Plain language",
        "Causal/logical",
        "Warning",
    ]
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(1, col, header), HEADER_FILL)
    field_labels = {spec.id: spec.label for spec in schema.fields}
    row_index = 2
    papers = {paper["id"]: paper for paper in store.list_papers()}
    warnings = store.list_warnings()
    warning_map: dict[tuple[str, str], list[str]] = {}
    for item in warnings:
        warning_map.setdefault((item["paper_id"], item["field_id"]), []).append(item["warning"])
    for item in store.list_evidence():
        extra = {}
        if item.get("extra_json"):
            try:
                extra = json.loads(item["extra_json"])
            except json.JSONDecodeError:
                extra = {}
        field_row = store.get_field(item["paper_id"], item["field_id"]) or {}
        values = [
            item["paper_id"],
            field_labels.get(item["field_id"], item["field_id"]),
            field_row.get("status"),
            field_row.get("confidence"),
            item.get("page"),
            item.get("quote"),
            "yes" if item.get("matched") else "no",
            item.get("score"),
            extra.get("academic_paraphrase"),
            extra.get("plain_language"),
            extra.get("causal_formulation"),
            "; ".join(warning_map.get((item["paper_id"], item["field_id"]), [])),
        ]
        fill = COL_FILLS["method"] if item.get("matched") else COL_FILLS["implications"]
        for col, value in enumerate(values, start=1):
            _style_body(ws.cell(row_index, col, value), fill)
        row_index += 1
    # warnings without evidence rows
    evidenced = {(item["paper_id"], item["field_id"]) for item in store.list_evidence()}
    for (paper_id, field_id), notes in warning_map.items():
        if (paper_id, field_id) in evidenced:
            continue
        field_row = store.get_field(paper_id, field_id) or {}
        values = [
            paper_id,
            field_labels.get(field_id, field_id),
            field_row.get("status"),
            field_row.get("confidence"),
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "; ".join(notes),
        ]
        for col, value in enumerate(values, start=1):
            _style_body(ws.cell(row_index, col, value), COL_FILLS["implications"])
        row_index += 1
        _ = papers
    for index, width in enumerate([22, 28, 14, 12, 8, 50, 14, 12, 36, 36, 36, 36], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(row_index - 1, 1)}"


def _run_log(wb: Workbook, store: Store) -> None:
    ws = wb.create_sheet("Run Log")
    headers = [
        "ID",
        "Paper ID",
        "Session",
        "Status",
        "Model",
        "Agent ID",
        "Run ID",
        "Schema",
        "Prompt",
        "Started",
        "Finished",
        "Error",
        "Cache key",
    ]
    for col, header in enumerate(headers, start=1):
        _style_header(ws.cell(1, col, header), HEADER_FILL)
    for row_index, item in enumerate(store.list_runs(), start=2):
        values = [
            item.get("id"),
            item.get("paper_id"),
            item.get("session_type"),
            item.get("status"),
            item.get("model"),
            item.get("agent_id"),
            item.get("run_id"),
            item.get("schema_version"),
            item.get("prompt_version"),
            item.get("started_at"),
            item.get("finished_at"),
            item.get("error"),
            item.get("cache_key"),
        ]
        for col, value in enumerate(values, start=1):
            _style_body(ws.cell(row_index, col, value), COL_FILLS["identity"])
    for index, width in enumerate([8, 22, 16, 12, 28, 22, 22, 12, 12, 22, 22, 40, 20], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def _notes(wb: Workbook) -> None:
    ws = wb.create_sheet("Notes")
    lines = [
        "How to use this workbook",
        "Each row on RARF Overview is one paper. Edit the 23 RARF columns as needed, then run `rarf sync-back` so your wording is stored as a human override and kept on the next export.",
        "Statuses: present = the paper reports it; not_reported = it could apply but the paper is silent; not_applicable = the paper type makes the field meaningless; unclear = the extractor could not decide.",
        "Framing.primary_basis is IV-led, DV-led, theory-led, or mixed/other. Style is theoretical vs phenomenological.",
        "Key argument cells keep an exact quotation plus academic, plain-language, and causal rephrasings. Check Evidence & QA for quote-match failures before trusting a quotation.",
        "Key variables are conceptual (nominal definitions). Measures are operationalizations. The Variable-Measure Map links them by construct_id.",
        "SQLite at data/rarf.sqlite is the source of truth. This workbook is the editable overview.",
        "LLM sessions use Cursor Grok 4.6 High. Cache keys combine PDF hash with schema, prompt, and model versions.",
    ]
    ws["A1"] = lines[0]
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    for index, line in enumerate(lines[1:], start=3):
        ws.cell(index, 1, line).font = ARIAL
        ws.cell(index, 1).alignment = WRAP
        ws.row_dimensions[index].height = 36
    ws.column_dimensions["A"].width = 110
