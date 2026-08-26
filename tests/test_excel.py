from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rarf_summarizer.excel_export import export_workbook
from rarf_summarizer.excel_import import sync_back
from rarf_summarizer.formatting import effective_text
from rarf_summarizer.schema import load_schema
from rarf_summarizer.storage import Store


def test_excel_round_trip_preserves_human_overrides(tmp_path: Path):
    store = Store(tmp_path / "rarf.sqlite")
    schema = load_schema()
    store.upsert_paper(
        {
            "id": "p1",
            "source_path": "x.pdf",
            "relative_path": "Duality/x.pdf",
            "folder": "Duality",
            "file_hash": "abc",
            "title": "CEO Duality",
            "authors": "Krause",
            "year": "2014",
            "doi": "10.1177/example",
            "page_count": 10,
            "warnings": "[]",
            "extracted_at": "2026-01-01T00:00:00+00:00",
            "status": "summarized",
        }
    )
    store.upsert_field(
        "p1",
        "research_question",
        {
            "status": "present",
            "confidence": 0.8,
            "generated_text": "What are the implications of CEO duality?",
            "generated_json": "{}",
        },
    )
    store.replace_constructs(
        "p1",
        [
            {
                "paper_id": "p1",
                "construct_id": "ceo_duality",
                "class": "IV",
                "name": "CEO duality",
                "nominal_definition": "CEO also serves as chair",
            }
        ],
    )
    store.replace_measures(
        "p1",
        [
            {
                "paper_id": "p1",
                "construct_id": "ceo_duality",
                "class": "IV",
                "name": "duality dummy",
                "operationalization": "1 if dual",
                "range": "0-1",
                "type": "binary",
                "linked_construct": "CEO duality",
            }
        ],
    )
    workbook = tmp_path / "RARF_Overview.xlsx"
    export_workbook(store, schema, workbook)
    wb = load_workbook(workbook)
    assert "RARF Overview" in wb.sheetnames
    assert "Variable-Measure Map" in wb.sheetnames
    overview = wb["RARF Overview"]
    headers = [overview.cell(2, col).value for col in range(1, overview.max_column + 1)]
    rq_col = headers.index("Research question") + 1
    overview.cell(3, rq_col).value = "Edited research question"
    edited = tmp_path / "edited.xlsx"
    wb.save(edited)
    updates = sync_back(store, schema, edited)
    assert updates == 1
    row = store.get_field("p1", "research_question")
    assert row["human_text"] == "Edited research question"
    assert effective_text(row) == "Edited research question"
    export_workbook(store, schema, workbook)
    again = load_workbook(workbook)
    headers = [again["RARF Overview"].cell(2, col).value for col in range(1, again["RARF Overview"].max_column + 1)]
    rq_col = headers.index("Research question") + 1
    assert again["RARF Overview"].cell(3, rq_col).value == "Edited research question"
    assert again["RARF Overview"].max_column == 6 + 23
