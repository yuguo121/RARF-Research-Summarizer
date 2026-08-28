from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from rarf_summarizer.excel_export import export_workbook, group_fills, overview_sheet_name
from rarf_summarizer.schema import apply_profile, load_schema
from rarf_summarizer.storage import Store
from rarf_summarizer.summarizer import session_names, session_section_keys, session_prompt
from rarf_summarizer.pdf_pipeline import ExtractedPaper

EXAMPLE = Path(__file__).resolve().parent.parent / "config" / "examples" / "quick_review.yaml"


def test_example_schema_loads_and_describes_itself():
    schema = load_schema(EXAMPLE)
    assert schema.name == "Quick Review"
    assert schema.name_short == "QR"
    assert len(schema.fields) == 5
    assert session_names(schema) == ["main"]
    assert "methods" in session_section_keys(schema, "main")
    groups = schema.groups_as_dict()
    assert [g["id"] for g in groups] == ["overview", "takeaways"]
    assert groups[0]["color"] == "#1F4E79"


def test_example_schema_prompt_uses_form_name_and_no_rarf_contracts():
    schema = load_schema(EXAMPLE)
    paper = ExtractedPaper(
        source_path=Path("x.pdf"),
        file_hash="abc",
        title="T",
        authors=None,
        doi=None,
        page_count=1,
        pages=[],
        sections=[],
    )
    prompt = session_prompt(schema, "main", "PACKET", paper)
    assert "Quick Review" in prompt
    assert "RARF" not in prompt
    assert "primary_basis" not in prompt  # no framing field, no framing contract


def test_example_schema_export_uses_dynamic_names_and_colors(tmp_path):
    schema = apply_profile(load_schema(EXAMPLE), None, None)
    store = Store(tmp_path / "test.sqlite")
    store.upsert_paper(
        {
            "id": "p1",
            "source_path": "x.pdf",
            "relative_path": "x.pdf",
            "folder": ".",
            "file_hash": "abc",
            "title": "Demo",
            "authors": "Author",
            "year": "2024",
            "doi": "",
            "page_count": 1,
            "warnings": "[]",
            "extracted_at": "2026-01-01T00:00:00+00:00",
            "status": "extracted",
        }
    )
    path = export_workbook(store, schema, tmp_path / "out.xlsx")
    wb = load_workbook(path)
    assert overview_sheet_name(schema) == "QR Overview"
    assert "QR Overview" in wb.sheetnames
    # No constructs/measures kinds in this schema -> no Variable-Measure Map sheet
    assert "Variable-Measure Map" not in wb.sheetnames
    strong, soft = group_fills(schema)
    assert "overview" in strong and "takeaways" in soft


def test_undeclared_group_gets_label_and_palette_fallback(tmp_path):
    raw = tmp_path / "schema.yaml"
    raw.write_text(
        """
name: "Minimal"
fields:
  - id: note
    label: Note
    group: misc_thoughts
    session: main
    value_kind: text
    instruction: Jot it down.
sessions:
  main:
    section_keys: [abstract]
""",
        encoding="utf-8",
    )
    schema = load_schema(raw)
    groups = schema.groups_as_dict()
    assert groups[0]["id"] == "misc_thoughts"
    assert groups[0]["label"] == "Misc Thoughts"
    strong, soft = group_fills(schema)
    assert "misc_thoughts" in strong and "misc_thoughts" in soft
