from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from rarf_summarizer.cursor_runtime import FakeBackend
from rarf_summarizer.pdf_pipeline import extract_paper
from rarf_summarizer.pipeline import Pipeline
from rarf_summarizer.summarizer import Summarizer
from tests.helpers import json_response, method_payload, theory_payload


def main() -> None:
    folder = Path(
        r"C:\Users\HuGoL\OneDrive - HKUST (Guangzhou)\PhD\Zotero\Corporate Governance\Internal"
    )
    pipe = Pipeline()
    papers = pipe.store.list_papers()
    if not papers:
        ids = pipe.extract_folder(folder)
        papers = pipe.store.list_papers()
    else:
        ids = [paper["id"] for paper in papers]
    krause = next(
        paper
        for paper in papers
        if "Krause, Semadeni, Cannella_2014" in (paper.get("relative_path") or paper.get("id") or "")
    )
    paper_id = krause["id"]
    extracted = extract_paper(Path(krause["source_path"]))
    pages = extracted.page_map()
    blob = pages.get(2, "") + " " + pages.get(3, "")
    match = re.search(r".{20}CEO duality.{40,120}", blob, re.IGNORECASE)
    quote = (match.group(0).strip() if match else blob.replace("\n", " ")[80:220]).strip()
    if len(quote) < 24:
        quote = "CEO duality—the practice of a single individual serving as both CEO and board chair"
    theory = theory_payload()
    theory["key_argument"]["value"][0]["quote"] = quote
    theory["key_argument"]["value"][0]["page"] = 2
    snippet = quote[:80]
    for key, payload in theory.items():
        if isinstance(payload, dict):
            payload["evidence"] = [{"page": 2, "quote": snippet}]
    method = method_payload()
    method_snippet = "48 publications"
    for key, payload in method.items():
        if isinstance(payload, dict) and payload.get("status") == "present":
            payload["evidence"] = [{"page": 3, "quote": method_snippet}]
    reconcile = {**theory, **method}
    backend = FakeBackend(
        {
            "theory": json_response(theory),
            "method": json_response(method),
            "reconcile": json_response(reconcile),
        }
    )
    Summarizer(pipe.store, pipe.schema, backend, pipe.work_dir).summarize_paper(
        paper_id, extracted, force=True
    )
    path = pipe.export()
    wb = load_workbook(path)
    overview = wb["RARF Overview"]
    headers = [overview.cell(2, col).value for col in range(1, overview.max_column + 1)]
    print("extracted_ids", ids)
    print("pages", extracted.page_count)
    print("doi", extracted.doi)
    print("sections", [(s.key, s.page_start) for s in extracted.sections[:15]])
    print("workbook", path)
    print("sheets", wb.sheetnames)
    print("overview_shape", overview.max_row, overview.max_column)
    print("header_count", len(headers))
    krause_row = None
    for row in range(3, overview.max_row + 1):
        if "Krause, Semadeni, Cannella_2014" in str(overview.cell(row, 6).value or ""):
            krause_row = row
            break
    krause_row = krause_row or 3
    print("krause_row", krause_row)
    print("framing", overview.cell(krause_row, headers.index("Framing") + 1).value)
    print("hypotheses", overview.cell(krause_row, headers.index("Hypotheses") + 1).value)
    print("measures", overview.cell(krause_row, headers.index("Measures") + 1).value)
    print("argument", str(overview.cell(krause_row, headers.index("Key argument") + 1).value)[:300])
    print("var_map_rows", wb["Variable-Measure Map"].max_row)
    print("evidence_rows", wb["Evidence & QA"].max_row)
    print("run_log_rows", wb["Run Log"].max_row)


if __name__ == "__main__":
    main()
